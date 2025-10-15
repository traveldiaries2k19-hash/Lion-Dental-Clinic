
import streamlit as st
from openpyxl import load_workbook

# Load the Excel workbook
excel_path = "Dental_Clinic_Template_Protected_Sheets.xlsx"
wb = load_workbook(excel_path)

# Function to append data to a sheet
def append_to_sheet(sheet_name, data):
    ws = wb[sheet_name]
    ws.append(data)
    wb.save(excel_path)

# Streamlit UI
st.title("Lion Dental Clinic And Implant Centre")
menu = st.sidebar.selectbox("Select Data Type", ["Patient Records", "Clinic Expenses", "Employee Salaries", "Upcoming Appointments"])

if menu == "Patient Records":
    st.header("Add Patient Record")
    date = st.date_input("Date")
    name = st.text_input("Patient Name")
    treatment = st.selectbox("Treatment", ["RCT", "Cleaning"])
    amount = st.number_input("Amount Charged", min_value=0)
    payment = st.selectbox("Payment Method", ["Cash", "Card", "UPI", "Insurance"])
    contact = st.text_input("Contact Number")
    notes = st.text_area("Special Notes")
    follow_up = st.date_input("Follow-up Date")
    if st.button("Submit Patient Record"):
        append_to_sheet("Patient Records", [date.strftime("%d/%m/%Y"), name, treatment, amount, payment, contact, notes, follow_up.strftime("%d/%m/%Y")])
        st.success("Patient record added successfully.")

elif menu == "Clinic Expenses":
    st.header("Add Clinic Expense")
    date = st.date_input("Date")
    item = st.text_input("Material/Equipment")
    vendor = st.text_input("Vendor")
    amount = st.number_input("Amount Spent", min_value=0.0)
    if st.button("Submit Expense"):
        append_to_sheet("Clinic Expenses", [date.strftime("%d/%m/%Y"), item, vendor, amount])
        st.success("Expense added successfully.")

elif menu == "Employee Salaries":
    st.header("Add Employee Salary")
    month = st.selectbox("Month", ["January", "February", "March", "April", "May", "June", "July", "August", "September", "October", "November", "December"])
    name = st.text_input("Employee Name")
    role = st.text_input("Role")
    amount = st.number_input("Amount Paid", min_value=0.0)
    mode = st.selectbox("Payment Mode", ["Bank Transfer", "Cash", "Cheque"])
    if st.button("Submit Salary"):
        append_to_sheet("Employee Salaries", [month, name, role, amount, mode])
        st.success("Salary record added successfully.")

elif menu == "Upcoming Appointments":
    st.header("Add Appointment")
    name = st.text_input("Patient Name")
    contact = st.text_input("Contact Number")
    treatment = st.text_input("Treatment")
    date = st.date_input("Date")
    time = st.text_input("Time")
    if st.button("Submit Appointment"):
        append_to_sheet("Upcoming Appointments", [name, contact, treatment, date.strftime("%Y-%m-%d"), time])
        st.success("Appointment added successfully.")


from openpyxl import load_workbook

def append_to_sheet(sheet_name, data):
    wb = load_workbook(excel_path)
    ws = wb[sheet_name]

    # Temporarily disable protection if needed (assuming no password)
    if ws.protection.sheet:
        ws.protection.sheet = False



if menu == "Patient Records":
    # Existing form code...

    st.subheader("Submitted Patient Records")
    try:
        df = Dental_Clinic_Template_Protected_Sheets.xlsx("Patient Records")
        st.dataframe(df)
    except Exception as e:
        st.error(f"Could not load data: {e}")


elif menu == "Clinic Expenses":
    # Existing form code...

    st.subheader("Submitted Clinic Expenses")
    try:
        df = load_sheet_data("Clinic Expenses")
        st.dataframe(df)
    except Exception as e:
        st.error(f"Could not load data: {e}")


elif menu == "Employee Salaries":
    # Existing form code...

    st.subheader("Submitted Employee Salaries")
    try:
        df = load_sheet_data("Employee Salaries")
        st.dataframe(df)
    except Exception as e:
        st.error(f"Could not load data: {e}")


def load_sheet_data(sheet_name):
    wb = load_workbook(excel_path)
    ws = wb[sheet_name]
    data = list(ws.values)
    headers = data[0]
    rows = data[1:]
    df = pd.DataFrame(rows, columns=headers)
    return df
